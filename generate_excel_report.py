import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def create_excel_report():
    """Создание Excel отчета из результатов тестов"""
    
    # Загружаем результаты тестов
    try:
        with open('test_results.json', 'r') as f:
            results = json.load(f)
    except FileNotFoundError:
        print("Результаты тестов не найдены")
        return
    
    # Создаем рабочую книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты тестов"
    
    # Заголовки
    headers = ['Название теста', 'Команда', 'Статус', 'Ожидаемый результат', 'Фактический результат', 'Время выполнения']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Данные
    for row, result in enumerate(results, 2):
        ws.cell(row=row, column=1, value=result['test'])
        ws.cell(row=row, column=2, value=result['command'])
        
        status_cell = ws.cell(row=row, column=3, value='ПРОЙДЕН' if result['success'] else 'НЕ ПРОЙДЕН')
        status_cell.fill = PatternFill(
            start_color="00FF00" if result['success'] else "FF0000",
            end_color="00FF00" if result['success'] else "FF0000", 
            fill_type="solid"
        )
        
        ws.cell(row=row, column=4, value=result['expected'])
        ws.cell(row=row, column=5, value=str(result['actual'])[:100])  # Обрезаем длинный вывод
        ws.cell(row=row, column=6, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    
    # Сводка
    summary_row = len(results) + 3
    passed = sum(1 for r in results if r['success'])
    failed = len(results) - passed
    
    ws.cell(row=summary_row, column=1, value="СВОДКА").font = Font(bold=True)
    ws.cell(row=summary_row+1, column=1, value=f"Всего тестов: {len(results)}")
    ws.cell(row=summary_row+2, column=1, value=f"Пройдено: {passed}")
    ws.cell(row=summary_row+3, column=1, value=f"Не пройдено: {failed}")
    ws.cell(row=summary_row+4, column=1, value=f"Процент успеха: {passed/len(results)*100:.1f}%")
    
    # Автоматическая настройка ширины колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем
    wb.save('test_report.xlsx')
    print(f"Excel отчет создан: test_report.xlsx")
    print(f"Результаты: {passed} пройдено, {failed} не пройдено из {len(results)} тестов")

if __name__ == "__main__":
    create_excel_report()