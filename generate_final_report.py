import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def generate_final_report():
    """Генерация итогового комплексного отчета тестирования"""
    
    # Загружаем результаты тестов
    test_results = []
    try:
        with open('test_results.json', 'r') as f:
            test_results = json.load(f)
        print("✓ test_results.json успешно загружен")
    except FileNotFoundError:
        print("ПРЕДУПРЕЖДЕНИЕ: test_results.json не найден")
    except json.JSONDecodeError as e:
        print(f"ОШИБКА: test_results.json содержит невалидный JSON: {e}")
    
    # Загружаем результаты сравнения - более надежная обработка
    comparison_results = {'summary': {'total_comparisons': 0, 'matches': 0, 'mismatches': 0, 'match_percentage': 0}, 'comparisons': []}
    
    comparison_file_path = 'comparison_results.json'
    if os.path.exists(comparison_file_path):
        file_size = os.path.getsize(comparison_file_path)
        if file_size == 0:
            print("ПРЕДУПРЕЖДЕНИЕ: comparison_results.json существует но пустой (0 байт)")
            print("Используются пустые результаты сравнения по умолчанию")
        else:
            try:
                with open(comparison_file_path, 'r') as f:
                    content = f.read().strip()
                    if content:
                        comparison_results = json.loads(content)
                        print("✓ comparison_results.json успешно загружен")
                    else:
                        print("ПРЕДУПРЕЖДЕНИЕ: comparison_results.json пустой после очистки")
            except json.JSONDecodeError as e:
                print(f"ОШИБКА: comparison_results.json содержит невалидный JSON: {e}")
                print("Используются пустые результаты сравнения по умолчанию")
    else:
        print("ПРЕДУПРЕЖДЕНИЕ: comparison_results.json не найден")
    
    # Создаем рабочую книгу с несколькими листами
    wb = Workbook()
    
    # Лист 1: Общая сводка
    ws_summary = wb.active
    ws_summary.title = "Общая сводка"
    
    # Сводка результатов тестов
    test_passed = sum(1 for r in test_results if r.get('success', False))
    test_total = len(test_results)
    test_success_rate = (test_passed / test_total * 100) if test_total > 0 else 0
    
    # Сводка сравнения
    comp_summary = comparison_results.get('summary', {})
    comp_match_rate = comp_summary.get('match_percentage', 0)
    
    # Общая сводка
    ws_summary.cell(1, 1, "ИТОГОВЫЙ ОТЧЕТ ТЕСТИРОВАНИЯ").font = Font(bold=True, size=16)
    ws_summary.cell(3, 1, "Сводка выполнения тестов").font = Font(bold=True)
    ws_summary.cell(4, 1, f"Всего тестов: {test_total}")
    ws_summary.cell(5, 1, f"Тестов пройдено: {test_passed}")
    ws_summary.cell(6, 1, f"Тестов не пройдено: {test_total - test_passed}")
    ws_summary.cell(7, 1, f"Процент успеха тестов: {test_success_rate:.1f}%")
    
    ws_summary.cell(9, 1, "Сводка сравнения изображений").font = Font(bold=True)
    ws_summary.cell(10, 1, f"Всего сравнений: {comp_summary.get('total_comparisons', 0)}")
    ws_summary.cell(11, 1, f"Совпадений изображений: {comp_summary.get('matches', 0)}")
    ws_summary.cell(12, 1, f"Несовпадений изображений: {comp_summary.get('mismatches', 0)}")
    ws_summary.cell(13, 1, f"Процент совпадения: {comp_match_rate:.1f}%")
    
    ws_summary.cell(17, 1, f"Отчет создан: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Лист 2: Детальные результаты тестов
    ws_tests = wb.create_sheet("Результаты тестов")
    
    headers = ['Название теста', 'Команда', 'Статус', 'Ожидаемый результат', 'Фактический результат']
    for col, header in enumerate(headers, 1):
        cell = ws_tests.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    for row, result in enumerate(test_results, 2):
        ws_tests.cell(row=row, column=1, value=result.get('test', 'Н/Д'))
        ws_tests.cell(row=row, column=2, value=result.get('command', 'Н/Д'))
        
        success = result.get('success', False)
        status_cell = ws_tests.cell(row=row, column=3, value='ПРОЙДЕН' if success else 'НЕ ПРОЙДЕН')
        status_cell.fill = PatternFill(
            start_color="00FF00" if success else "FF0000",
            end_color="00FF00" if success else "FF0000", 
            fill_type="solid"
        )
        
        ws_tests.cell(row=row, column=4, value=result.get('expected', 'Н/Д'))
        ws_tests.cell(row=row, column=5, value=str(result.get('actual', 'Н/Д'))[:100])
    
    # Лист 3: Результаты сравнения изображений
    ws_compare = wb.create_sheet("Сравнение изображений")
    
    headers = ['Эталонный файл', 'Выходной файл', 'Совпадение', 'Размер эталона', 'Размер выхода', 'Метод']
    for col, header in enumerate(headers, 1):
        cell = ws_compare.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    comparisons = comparison_results.get('comparisons', [])
    if comparisons:
        for row, comp in enumerate(comparisons, 2):
            ws_compare.cell(row=row, column=1, value=comp.get('standard_file', 'Н/Д'))
            ws_compare.cell(row=row, column=2, value=comp.get('output_file', 'Н/Д'))
            
            match = comp.get('match', False)
            match_cell = ws_compare.cell(row=row, column=3, value='ДА' if match else 'НЕТ')
            match_cell.fill = PatternFill(
                start_color="00FF00" if match else "FF0000",
                end_color="00FF00" if match else "FF0000", 
                fill_type="solid"
            )
            
            ws_compare.cell(row=row, column=4, value=comp.get('file_size_std', 'Н/Д'))
            ws_compare.cell(row=row, column=5, value=comp.get('file_size_out', 'Н/Д'))
            ws_compare.cell(row=row, column=6, value=comp.get('method', 'Н/Д'))
    else:
        # Если нет данных сравнения, добавить сообщение
        ws_compare.cell(2, 1, "Данные сравнения недоступны")
        ws_compare.merge_cells('A2:F2')
        ws_compare.cell(2, 1).alignment = Alignment(horizontal='center')
    
    # Автоматическая настройка ширины колонок для всех листов
    for ws in [ws_summary, ws_tests, ws_compare]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем итоговый отчет
    report_filename = 'final_test_report.xlsx'
    wb.save(report_filename)
    print(f"✓ Итоговый отчет тестирования создан: {report_filename}")
    
    # Выводим сводку в консоль
    print(f"\n=== ИТОГОВАЯ СВОДКА ТЕСТИРОВАНИЯ ===")
    print(f"Тесты: {test_passed}/{test_total} пройдено ({test_success_rate:.1f}%)")
    print(f"Изображения: {comp_summary.get('matches', 0)}/{comp_summary.get('total_comparisons', 0)} совпало ({comp_match_rate:.1f}%)")

def fix_comparison_results_file():
    """Создает валидный пустой файл comparison_results.json если он пустой или отсутствует"""
    comparison_file_path = 'comparison_results.json'
    
    if os.path.exists(comparison_file_path):
        file_size = os.path.getsize(comparison_file_path)
        if file_size == 0:
            print("Исправляем пустой файл comparison_results.json...")
            # Создаем валидный пустой JSON
            empty_data = {
                "summary": {
                    "total_comparisons": 0,
                    "matches": 0, 
                    "mismatches": 0,
                    "match_percentage": 0
                },
                "comparisons": []
            }
            with open(comparison_file_path, 'w') as f:
                json.dump(empty_data, f, indent=2)
            print("✓ Создан валидный пустой comparison_results.json")
        else:
            print("✓ comparison_results.json уже существует и не пустой")
    else:
        print("Создаем отсутствующий файл comparison_results.json...")
        # Создаем файл с валидной структурой
        empty_data = {
            "summary": {
                "total_comparisons": 0,
                "matches": 0,
                "mismatches": 0,
                "match_percentage": 0
            },
            "comparisons": []
        }
        with open(comparison_file_path, 'w') as f:
            json.dump(empty_data, f, indent=2)
        print("✓ Создан comparison_results.json с валидной структурой")

if __name__ == "__main__":
    # Сначала исправляем файл comparison_results.json если нужно
    print("Проверяем файл comparison_results.json...")
    fix_comparison_results_file()
    
    print("\n" + "="*50)
    print("Генерируем итоговый отчет...")
    print("="*50 + "\n")
    
    # Затем генерируем отчет
    generate_final_report()